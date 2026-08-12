"""
Build the competitor messaging matrix from a sheet of ads.

Reads a CSV of collected ads, uses Claude to tag each ad by message theme,
audience, and proof points, then writes a formatted Excel workbook: the
strategic findings, the claim matrix, the audience matrix, recurring proof
points, and the tagged detail behind it all.

Two models, each doing what it is best at:
  - Haiku tags each ad (simple, high-volume extraction)
  - Opus reads the finished matrix and writes the strategic findings

Usage:
    venv/bin/python src/build_matrix.py [path/to/ads.csv]

Default input:  data/sample_ads.csv
Output:         output/messaging_matrix.xlsx
"""
import csv
import hashlib
import json
import os
import sys
from collections import defaultdict

from dotenv import load_dotenv
from anthropic import Anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# EDIT THESE TWO LISTS to fit your category. They become the matrix columns.
# ---------------------------------------------------------------------------

THEMES = [
    "Performance / Technical fabric",
    "Comfort / Feel",
    "Style / Design",
    "Versatility / Gym-to-street",
    "Sustainability",
    "Price / Value",
    "Innovation / New technology",
    "Community / Belonging",
]

AUDIENCES = [
    "Women",
    "Men",
    "Runners",
    "Yoga & studio",
    "Gym & strength training",
    "Outdoor & hiking",
    "Everyday / athleisure",
    "Size-inclusive",
]

# What kind of ad this is. Search results are dominated by product-feed and
# promotional ads, which say nothing about how a brand positions itself.
# Tagging the type lets the matrix be run on brand-level ads alone.
AD_TYPES = [
    "Brand",          # says what the brand stands for
    "Product",        # pushes a specific item or category
    "Promotional",    # led by a discount, code, sale or loyalty offer
]

# ---------------------------------------------------------------------------

TAG_MODEL = "claude-haiku-4-5"      # tags each ad: cheap, fast, accurate enough
INSIGHT_MODEL = "claude-opus-5"     # reads the matrix and writes the findings

# Resolve paths from the project root, so the script runs from anywhere.
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

load_dotenv(os.path.join(BASE, ".env"))
API_KEY = os.getenv("ANTHROPIC_API_KEY")
if not API_KEY:
    print("No API key found. Paste it into .env, then run check_key.py.")
    raise SystemExit(1)

client = Anthropic(api_key=API_KEY)

REQUIRED_COLUMNS = ["competitor", "ad_text", "format", "first_seen"]

TAG_SCHEMA = {
    "type": "object",
    "properties": {
        "themes": {
            "type": "array",
            "items": {"type": "string", "enum": THEMES},
        },
        "audiences": {
            "type": "array",
            "items": {"type": "string", "enum": AUDIENCES},
        },
        "ad_type": {"type": "string", "enum": AD_TYPES},
        "proof_points": {"type": "array", "items": {"type": "string"}},
    },
    "required": ["themes", "audiences", "ad_type", "proof_points"],
    "additionalProperties": False,
}

FINDINGS_SCHEMA = {
    "type": "object",
    "properties": {
        "headline": {"type": "string"},
        "whitespace": {
            "type": "object",
            "properties": {
                "opportunity": {"type": "string"},
                "why_it_matters": {"type": "string"},
                "recommended_move": {"type": "string"},
            },
            "required": ["opportunity", "why_it_matters", "recommended_move"],
            "additionalProperties": False,
        },
        "competitors": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                    "positioning": {"type": "string"},
                    "who_they_target": {"type": "string"},
                    "signature_proof": {"type": "string"},
                },
                "required": ["name", "positioning", "who_they_target", "signature_proof"],
                "additionalProperties": False,
            },
        },
        "recurring_proof_points": {"type": "array", "items": {"type": "string"}},
        "recommendations": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "recommendation": {"type": "string"},
                    "rationale": {"type": "string"},
                },
                "required": ["recommendation", "rationale"],
                "additionalProperties": False,
            },
        },
    },
    "required": [
        "headline",
        "whitespace",
        "competitors",
        "recurring_proof_points",
        "recommendations",
    ],
    "additionalProperties": False,
}


# ---------------------------------------------------------------------------
# Input validation
# ---------------------------------------------------------------------------

def read_ads(path):
    if not os.path.exists(path):
        print(f"No file at {path}")
        print("Put the team's collected ads there as a CSV, or pass a path:")
        print("  venv/bin/python src/build_matrix.py data/your_ads.csv")
        raise SystemExit(1)
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def validate(rows, path):
    """Check the sheet before spending time and tokens on a bad run."""
    name = os.path.basename(path)

    if not rows:
        print(f"{name} has a header but no ad rows.")
        raise SystemExit(1)

    missing = [c for c in REQUIRED_COLUMNS if c not in rows[0]]
    if missing:
        print(f"{name} is missing required column(s): {', '.join(missing)}")
        print(f"The sheet needs these exact columns: {', '.join(REQUIRED_COLUMNS)}")
        raise SystemExit(1)

    usable, warnings = [], []
    for i, r in enumerate(rows, start=2):  # row 1 is the header
        if not (r.get("competitor") or "").strip():
            warnings.append(f"  row {i}: no competitor name, skipped")
            continue
        if not (r.get("ad_text") or "").strip():
            warnings.append(f"  row {i}: no ad text, skipped")
            continue
        usable.append(r)

    seen = set()
    duplicates = 0
    deduped = []
    for r in usable:
        key = (r["competitor"].strip().lower(), " ".join(r["ad_text"].split()).lower())
        if key in seen:
            duplicates += 1
            continue
        seen.add(key)
        deduped.append(r)

    if warnings:
        print(f"Sheet warnings in {name}:")
        for w in warnings:
            print(w)
    if duplicates:
        print(f"  removed {duplicates} duplicate ad(s)")

    if not deduped:
        print(f"{name} has no usable ads after checks.")
        raise SystemExit(1)

    return deduped


# ---------------------------------------------------------------------------
# Model calls
# ---------------------------------------------------------------------------

def tag_ad(competitor, ad_text):
    """Tag one ad. Returns {themes, audiences, proof_points}."""
    prompt = (
        "You are analyzing a competitor's advertisement.\n\n"
        f"Brand: {competitor}\n"
        f'Ad copy: "{ad_text}"\n\n'
        "Choose every message theme the ad clearly uses, every audience it "
        "clearly speaks to, and any concrete proof points it cites "
        "(numbers, customer counts, guarantees, awards, ratings). "
        "An ad can address more than one audience: name each one the copy "
        "actually signals, not just the broadest. If the ad cites no concrete "
        "proof, return an empty list.\n\n"
        "Also classify the ad. Brand means the copy says what the brand stands "
        "for or who it is for. Product means it pushes a specific item or "
        "category with no wider claim. Promotional means a discount, code, "
        "sale or loyalty offer is doing the work."
    )
    resp = client.messages.create(
        model=TAG_MODEL,
        max_tokens=500,
        output_config={"format": {"type": "json_schema", "schema": TAG_SCHEMA}},
        messages=[{"role": "user", "content": prompt}],
    )
    text = next((b.text for b in resp.content if b.type == "text"), "{}")
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        data = {}
    return {
        # dict.fromkeys dedupes while keeping the model's ordering, so one ad
        # can never count twice against the same theme.
        "themes": list(dict.fromkeys(t for t in data.get("themes", []) if t in THEMES)),
        "audiences": list(
            dict.fromkeys(a for a in data.get("audiences", []) if a in AUDIENCES)
        ),
        "ad_type": data.get("ad_type") if data.get("ad_type") in AD_TYPES else "Product",
        "proof_points": data.get("proof_points", []) or [],
    }


# ---------------------------------------------------------------------------
# Frozen tags
#
# The model does not return identical tags for identical copy on every run, so
# rebuilding from scratch each time makes the numbers move under you. Tags are
# written to data/tags.json, keyed by a hash of the ad, and reused on later
# runs. The file is committed, so the matrix is reproducible and a reviewer's
# correction to a tag survives the next rebuild. Delete the file, or pass
# --retag, to score everything again from scratch.
# ---------------------------------------------------------------------------

TAGS_PATH = os.path.join(BASE, "data", "tags.json")


def ad_key(competitor, ad_text):
    """Stable id for one ad, so tags survive reordering of the sheet."""
    return hashlib.sha256(f"{competitor}\x00{ad_text}".encode("utf-8")).hexdigest()[:16]


def load_frozen_tags():
    if not os.path.exists(TAGS_PATH):
        return {}
    try:
        with open(TAGS_PATH, encoding="utf-8") as fh:
            return json.load(fh).get("tags", {})
    except (json.JSONDecodeError, OSError):
        print("  tags.json unreadable, re-tagging everything")
        return {}


def save_frozen_tags(frozen):
    os.makedirs(os.path.dirname(TAGS_PATH), exist_ok=True)
    payload = {
        "_readme": (
            "Frozen ad tags. Edit a value here to correct a mis-tagged ad and the "
            "correction survives future runs. Delete this file, or run with "
            "--retag, to score every ad again from scratch."
        ),
        "themes": THEMES,
        "audiences": AUDIENCES,
        "ad_types": AD_TYPES,
        "tags": frozen,
    }
    with open(TAGS_PATH, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False, sort_keys=True)
        fh.write("\n")


def write_findings(rows, tags, theme_counts, audience_counts, proof_by_comp, whitespace):
    """Ask Opus to read the finished matrix and write the strategic take."""
    competitors = sorted({r["competitor"] for r in rows})

    lines = ["CLAIM MATRIX (ads per theme):"]
    for c in competitors:
        used = [f"{t} ({theme_counts[c][t]})" for t in THEMES if theme_counts[c][t]]
        lines.append(f"  {c}: {', '.join(used) if used else 'none'}")

    lines.append("")
    lines.append("AUDIENCE MATRIX (ads per audience):")
    for c in competitors:
        used = [f"{a} ({audience_counts[c][a]})" for a in AUDIENCES if audience_counts[c][a]]
        lines.append(f"  {c}: {', '.join(used) if used else 'none'}")

    lines.append("")
    lines.append("PROOF POINTS CITED:")
    for c in competitors:
        pts = proof_by_comp.get(c, [])
        lines.append(f"  {c}: {'; '.join(pts) if pts else 'none'}")

    lines.append("")
    lines.append(f"THEMES NO COMPETITOR USES: {', '.join(whitespace) if whitespace else 'none'}")

    lines.append("")
    lines.append("EXAMPLE AD COPY (up to 3 per competitor):")
    shown = defaultdict(int)
    for r in rows:
        c = r["competitor"]
        if shown[c] >= 3:
            continue
        shown[c] += 1
        lines.append(f'  {c}: "{r["ad_text"]}"')

    prompt = (
        "You are a marketing strategist analyzing what a set of competitors are "
        "advertising right now. Below is a messaging matrix built from their live ads.\n\n"
        + "\n".join(lines)
        + "\n\nWrite the strategic read on this data:\n"
        "- headline: one sentence capturing the single most important finding.\n"
        "- whitespace: the clearest unclaimed angle. Say what it is, why it matters "
        "commercially, and the concrete move to make. If every theme is contested, "
        "identify the least crowded one and say so plainly.\n"
        "- competitors: for each brand, a one-line positioning summary, who they target, "
        "and their signature proof point.\n"
        "- recurring_proof_points: the proof types that repeat across the category.\n"
        "- recommendations: two or three specific actions, each with its rationale.\n\n"
        "Write in plain business English for a marketing leader. Be concrete and "
        "decision-oriented. No hedging, no filler."
    )

    kwargs = dict(
        model=INSIGHT_MODEL,
        max_tokens=16000,
        output_config={"format": {"type": "json_schema", "schema": FINDINGS_SCHEMA}},
        messages=[{"role": "user", "content": prompt}],
    )

    # Server-side fallback keeps the run alive if a safety classifier declines.
    # If the beta is unavailable on this account, fall back to the standard call.
    try:
        with client.beta.messages.stream(
            betas=["server-side-fallback-2026-07-01"],
            fallbacks="default",
            **kwargs,
        ) as stream:
            resp = stream.get_final_message()
    except Exception:
        with client.messages.stream(**kwargs) as stream:
            resp = stream.get_final_message()

    if resp.stop_reason == "refusal":
        print("  the model declined to analyze this data; skipping the findings sheet")
        return None

    text = next((b.text for b in resp.content if b.type == "text"), None)
    if not text:
        return None
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return None


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WHITESPACE_FILL = PatternFill("solid", fgColor="FCE4D6")
IN_PLAY_FILL = PatternFill("solid", fgColor="DDEBF7")


def _header_row(ws, labels, row=1):
    for j, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=j, value=label)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER


def _count_matrix(ws, competitors, columns, counts, highlight_empty, ads_per_comp):
    """Shared layout for the claim and audience matrices. Returns empty columns.

    Cells hold the percent of that brand's ads using the column, not the raw
    count. Brands are sampled unevenly (one may have 18 ads collected and
    another 3), so raw counts would rank brands by how much of them we happened
    to collect rather than by what they actually say. Column B keeps the
    denominator visible so nobody reads a percent without knowing the base.
    """
    _header_row(ws, ["Competitor", "Ads"] + columns)
    first_data_col = 3

    for i, comp in enumerate(competitors, start=2):
        bc = ws.cell(row=i, column=1, value=comp)
        bc.font = BOLD
        bc.alignment = LEFT
        bc.border = BORDER
        n_ads = ads_per_comp[comp]
        nc = ws.cell(row=i, column=2, value=n_ads)
        nc.alignment = CENTER
        nc.border = BORDER
        for j, col in enumerate(columns, start=first_data_col):
            n = counts[comp][col]
            pct = round(100 * n / n_ads) if n_ads else 0
            c = ws.cell(row=i, column=j, value=(pct if n else None))
            c.number_format = '0"%"'
            c.alignment = CENTER
            c.border = BORDER
            if n:
                c.fill = IN_PLAY_FILL

    total_row = len(competitors) + 2
    tc = ws.cell(row=total_row, column=1, value="CATEGORY %")
    tc.font = BOLD
    tc.alignment = LEFT
    tc.border = BORDER
    last = total_row - 1
    nc = ws.cell(row=total_row, column=2, value=f"=SUM(B2:B{last})")
    nc.font = BOLD
    nc.alignment = CENTER
    nc.border = BORDER
    for j in range(first_data_col, len(columns) + first_data_col):
        letter = get_column_letter(j)
        c = ws.cell(row=total_row, column=j)
        # Weight each brand's percent by its ad count to recover the true
        # share of all ads, so the row stays live if a cell is edited.
        c.value = (
            f"=ROUND(SUMPRODUCT($B$2:$B${last},{letter}2:{letter}{last})"
            f"/SUM($B$2:$B${last}),0)"
        )
        c.number_format = '0"%"'
        c.font = BOLD
        c.alignment = CENTER
        c.border = BORDER

    empty = []
    for j, col in enumerate(columns, start=first_data_col):
        if sum(counts[comp][col] for comp in competitors) == 0:
            empty.append(col)
            if highlight_empty:
                for i in range(1, total_row + 1):
                    ws.cell(row=i, column=j).fill = WHITESPACE_FILL

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 7
    for j in range(first_data_col, len(columns) + first_data_col):
        ws.column_dimensions[get_column_letter(j)].width = 15
    ws.freeze_panes = "C2"
    return empty, total_row


def _findings_sheet(ws, findings):
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 100

    row = 1
    ws.cell(row=row, column=1, value="THE FINDING").font = Font(bold=True, color="1F3864", size=12)
    hc = ws.cell(row=row, column=2, value=findings["headline"])
    hc.font = Font(bold=True, size=12)
    hc.alignment = LEFT
    ws.row_dimensions[row].height = 32
    row += 2

    ws.cell(row=row, column=1, value="WHITESPACE").font = Font(bold=True, color="C55A11", size=12)
    row += 1
    for label, key in [
        ("The opportunity", "opportunity"),
        ("Why it matters", "why_it_matters"),
        ("Recommended move", "recommended_move"),
    ]:
        ws.cell(row=row, column=1, value=label).font = BOLD
        ws.cell(row=row, column=1).alignment = LEFT
        c = ws.cell(row=row, column=2, value=findings["whitespace"][key])
        c.alignment = LEFT
        ws.row_dimensions[row].height = 45
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="HOW EACH BRAND IS POSITIONED").font = Font(
        bold=True, color="1F3864", size=12
    )
    row += 1
    for comp in findings["competitors"]:
        ws.cell(row=row, column=1, value=comp["name"]).font = BOLD
        ws.cell(row=row, column=1).alignment = LEFT
        body = (
            f"{comp['positioning']}\n"
            f"Targets: {comp['who_they_target']}\n"
            f"Signature proof: {comp['signature_proof']}"
        )
        c = ws.cell(row=row, column=2, value=body)
        c.alignment = LEFT
        ws.row_dimensions[row].height = 58
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="PROOF POINTS THE CATEGORY LEANS ON").font = Font(
        bold=True, color="1F3864", size=12
    )
    row += 1
    for pt in findings["recurring_proof_points"]:
        c = ws.cell(row=row, column=2, value=pt)
        c.alignment = LEFT
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="RECOMMENDATIONS").font = Font(
        bold=True, color="1F3864", size=12
    )
    row += 1
    for i, rec in enumerate(findings["recommendations"], start=1):
        ws.cell(row=row, column=1, value=f"{i}.").font = BOLD
        ws.cell(row=row, column=1).alignment = LEFT
        c = ws.cell(row=row, column=2, value=f"{rec['recommendation']}\nWhy: {rec['rationale']}")
        c.alignment = LEFT
        ws.row_dimensions[row].height = 48
        row += 1


def build_workbook(rows, tags, findings):
    competitors = sorted({r["competitor"] for r in rows})

    theme_counts = defaultdict(lambda: defaultdict(int))
    audience_counts = defaultdict(lambda: defaultdict(int))
    proof_counts = defaultdict(lambda: defaultdict(int))
    ads_per_comp = defaultdict(int)
    for r, t in zip(rows, tags):
        comp = r["competitor"]
        ads_per_comp[comp] += 1
        for theme in t["themes"]:
            theme_counts[comp][theme] += 1
        for audience in t["audiences"]:
            audience_counts[comp][audience] += 1
        for pt in t["proof_points"]:
            proof_counts[comp][pt.strip()] += 1

    wb = Workbook()

    # ---- Sheet 1: Findings (the analysis leads) ----
    ws_find = wb.active
    ws_find.title = "Findings"

    # ---- Sheet 2: Messaging Matrix (claims) ----
    ws = wb.create_sheet("Messaging Matrix")
    whitespace, total_row = _count_matrix(
        ws, competitors, THEMES, theme_counts, True, ads_per_comp
    )
    note = ws.cell(
        row=total_row + 2,
        column=1,
        value="Cell = percent of that brand's ads using the theme. Ads = how many of "
        "that brand's ads we sampled, which is the denominator. Percentages let "
        "brands with different sample sizes be compared. Blue = a theme in play. "
        "Peach columns = whitespace (no competitor is using it).",
    )
    note.font = Font(italic=True, color="666666")

    # ---- Sheet 3: Audience Matrix ----
    ws2 = wb.create_sheet("Audience Matrix")
    unaddressed, total_row2 = _count_matrix(
        ws2, competitors, AUDIENCES, audience_counts, True, ads_per_comp
    )
    note2 = ws2.cell(
        row=total_row2 + 2,
        column=1,
        value="Cell = percent of that brand's ads aimed at the audience. Ads = the "
        "denominator. Peach columns = audiences no competitor is speaking to.",
    )
    note2.font = Font(italic=True, color="666666")

    # ---- Sheet 4: Proof Points ----
    ws3 = wb.create_sheet("Proof Points")
    _header_row(ws3, ["Competitor", "Proof point cited", "Times used"])
    r_i = 2
    for comp in competitors:
        pts = sorted(proof_counts[comp].items(), key=lambda kv: (-kv[1], kv[0]))
        if not pts:
            for j, v in enumerate([comp, "(no concrete proof points cited)", 0], start=1):
                c = ws3.cell(row=r_i, column=j, value=v)
                c.alignment = LEFT if j <= 2 else CENTER
                c.border = BORDER
            ws3.cell(row=r_i, column=1).font = BOLD
            r_i += 1
            continue
        for k, (pt, n) in enumerate(pts):
            for j, v in enumerate([comp if k == 0 else "", pt, n], start=1):
                c = ws3.cell(row=r_i, column=j, value=v)
                c.alignment = LEFT if j <= 2 else CENTER
                c.border = BORDER
            if k == 0:
                ws3.cell(row=r_i, column=1).font = BOLD
            r_i += 1
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 70
    ws3.column_dimensions["C"].width = 12
    ws3.freeze_panes = "A2"

    # ---- Sheet 5: Tagged Ads (the detail behind the matrix) ----
    ws4 = wb.create_sheet("Tagged Ads")
    _header_row(
        ws4,
        ["Competitor", "Ad copy", "Ad type", "Themes", "Audience", "Proof points", "Format", "First seen"],
    )
    for i, (r, t) in enumerate(zip(rows, tags), start=2):
        vals = [
            r["competitor"],
            r["ad_text"],
            t["ad_type"],
            ", ".join(t["themes"]),
            ", ".join(t["audiences"]),
            ", ".join(t["proof_points"]),
            r.get("format", ""),
            r.get("first_seen", ""),
        ]
        for j, v in enumerate(vals, start=1):
            c = ws4.cell(row=i, column=j, value=v)
            c.alignment = LEFT
            c.border = BORDER
    for j, w in enumerate([16, 55, 28, 24, 30, 10, 12], start=1):
        ws4.column_dimensions[get_column_letter(j)].width = w
    ws4.freeze_panes = "A2"

    # Fill the Findings sheet last, now that everything else is computed.
    if findings:
        _findings_sheet(ws_find, findings)
    else:
        ws_find.column_dimensions["A"].width = 100
        ws_find.cell(
            row=1,
            column=1,
            value="Findings could not be generated on this run. "
            "The matrix sheets are complete; rerun to try again.",
        ).alignment = LEFT

    return wb, whitespace, unaddressed, proof_counts


# ---------------------------------------------------------------------------

def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    flags = {a for a in sys.argv[1:] if a.startswith("--")}
    retag = "--retag" in flags
    brand_only = "--brand-only" in flags

    in_path = args[0] if args else os.path.join(BASE, "data", "sample_ads.csv")
    if not os.path.isabs(in_path):
        in_path = os.path.join(BASE, in_path)

    rows = validate(read_ads(in_path), in_path)
    print(f"Read {len(rows)} ads from {os.path.relpath(in_path, BASE)}.")

    frozen = {} if retag else load_frozen_tags()
    if retag:
        print("--retag: scoring every ad again, ignoring saved tags.")

    tags = []
    reused = 0
    for r in rows:
        key = ad_key(r["competitor"], r["ad_text"])
        t = frozen.get(key)
        if t and all(k in t for k in ("themes", "audiences", "ad_type", "proof_points")):
            reused += 1
        else:
            t = tag_ad(r["competitor"], r["ad_text"])
            frozen[key] = t
            print(f"  tagged {r['competitor']}: {', '.join(t['themes']) or '(no clear theme)'}")
        tags.append(t)

    print(f"\nReused {reused} saved tag(s), scored {len(rows) - reused} new ad(s).")
    save_frozen_tags(frozen)
    print(f"Tags frozen in {os.path.relpath(TAGS_PATH, BASE)}")

    if brand_only:
        keep = [(r, t) for r, t in zip(rows, tags) if t["ad_type"] == "Brand"]
        dropped = len(rows) - len(keep)
        if not keep:
            print("--brand-only left no ads. Run without the flag.")
            raise SystemExit(1)
        rows = [r for r, _ in keep]
        tags = [t for _, t in keep]
        print(f"--brand-only: kept {len(rows)} brand-level ad(s), set aside {dropped}.")

    mix = defaultdict(int)
    for t in tags:
        mix[t["ad_type"]] += 1
    print("Ad mix: " + ", ".join(f"{k} {v}" for k, v in sorted(mix.items())))

    # Precompute what the findings pass needs.
    competitors = sorted({r["competitor"] for r in rows})
    theme_counts = defaultdict(lambda: defaultdict(int))
    audience_counts = defaultdict(lambda: defaultdict(int))
    proof_by_comp = defaultdict(list)
    for r, t in zip(rows, tags):
        comp = r["competitor"]
        for theme in t["themes"]:
            theme_counts[comp][theme] += 1
        for audience in t["audiences"]:
            audience_counts[comp][audience] += 1
        for pt in t["proof_points"]:
            if pt.strip() and pt.strip() not in proof_by_comp[comp]:
                proof_by_comp[comp].append(pt.strip())

    whitespace_preview = [
        t for t in THEMES if sum(theme_counts[c][t] for c in competitors) == 0
    ]

    print("\nWriting the strategic findings...")
    findings = write_findings(
        rows, tags, theme_counts, audience_counts, proof_by_comp, whitespace_preview
    )

    wb, whitespace, unaddressed, _ = build_workbook(rows, tags, findings)
    out_dir = os.path.join(BASE, "output")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "messaging_matrix.xlsx")
    wb.save(out_path)

    print(f"\nSaved to {os.path.relpath(out_path, BASE)}")
    if findings:
        print(f"\nHeadline: {findings['headline']}")
        print(f"Whitespace: {findings['whitespace']['opportunity']}")
    if whitespace:
        print("\nClaim themes no competitor uses:")
        for w in whitespace:
            print(f"  - {w}")
    if unaddressed:
        print("Audiences no competitor targets:")
        for a in unaddressed:
            print(f"  - {a}")


if __name__ == "__main__":
    main()
