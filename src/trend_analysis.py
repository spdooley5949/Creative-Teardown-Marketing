"""
Longitudinal messaging analysis.

Reads a CSV of ads that carry real first_seen dates, tags each one with the
same themes and audiences the matrix uses, and reports how the category's
claims shifted year over year.

This is deliberately separate from build_matrix.py. That tool answers "what is
the category saying now." This one answers "what changed." They use the same
theme definitions but different datasets and different frozen tag files, so
neither can disturb the other.

Usage:
    venv/bin/python src/trend_analysis.py data/competitor_ads_5yr_v2.csv
    venv/bin/python src/trend_analysis.py <csv> --retag     # ignore saved tags
"""
import csv
import datetime
import hashlib
import json
import os
import sys
import threading
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from build_matrix import BASE, THEMES, tag_ad

TAGS_PATH = os.path.join(BASE, "data", "tags_trend.json")
OUT_PATH = os.path.join(BASE, "output", "messaging_trend.xlsx")
WORKERS = 12                       # tagging is IO-bound; keep well under rate limits
DATE_FORMATS = ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%d/%m/%Y")

INK = "FF14181C"
SIG = "FFE8562A"
HEAD_FILL = PatternFill("solid", fgColor="FF14181C")
SIG_FILL = PatternFill("solid", fgColor="FFFDEEE8")


def parse_year(raw):
    raw = (raw or "").strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.datetime.strptime(raw, fmt).year
        except ValueError:
            continue
    return None


def key(row):
    """Stable id for an ad, so tags survive between runs."""
    return hashlib.sha1(
        (row["competitor"] + "|" + row["ad_text"]).encode("utf-8")
    ).hexdigest()[:16]


def load_rows(path):
    with open(path, newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))
    required = {"competitor", "ad_text", "first_seen"}
    missing = required - set(rows[0])
    if missing:
        print(f"{path} is missing column(s): {', '.join(sorted(missing))}")
        raise SystemExit(1)

    kept, undated = [], 0
    for r in rows:
        year = parse_year(r["first_seen"])
        if year is None:
            undated += 1
            continue
        r["_year"] = year
        r["_key"] = key(r)
        kept.append(r)
    if undated:
        print(f"  skipped {undated} row(s) with an unreadable first_seen date")
    return kept


def tag_all(rows, retag):
    cache = {}
    if os.path.exists(TAGS_PATH) and not retag:
        with open(TAGS_PATH, encoding="utf-8") as fh:
            cache = json.load(fh)

    todo = [r for r in rows if r["_key"] not in cache]
    print(f"Reused {len(rows) - len(todo)} saved tag(s), scoring {len(todo)} new ad(s).")

    if todo:
        lock = threading.Lock()
        done = [0]

        def work(r):
            try:
                tags = tag_ad(r["competitor"], r["ad_text"])
            except Exception as exc:                      # keep one bad ad from killing the run
                tags = {"themes": [], "audiences": [], "proof_points": [], "_error": str(exc)}
            with lock:
                cache[r["_key"]] = tags
                done[0] += 1
                if done[0] % 100 == 0 or done[0] == len(todo):
                    print(f"  tagged {done[0]}/{len(todo)}")
            return tags

        with ThreadPoolExecutor(max_workers=WORKERS) as pool:
            list(pool.map(work, todo))

        os.makedirs(os.path.dirname(TAGS_PATH), exist_ok=True)
        with open(TAGS_PATH, "w", encoding="utf-8") as fh:
            json.dump(cache, fh, indent=1, sort_keys=True)
        print(f"Saved tags to {os.path.relpath(TAGS_PATH, BASE)}")

    failed = sum(1 for r in rows if cache.get(r["_key"], {}).get("_error"))
    if failed:
        print(f"  warning: {failed} ad(s) failed to tag and are counted as untagged")
    return cache


def build(rows, cache):
    years = sorted({r["_year"] for r in rows})
    per_year = Counter(r["_year"] for r in rows)
    theme_year = defaultdict(Counter)
    brand_year = defaultdict(Counter)
    for r in rows:
        tags = cache.get(r["_key"], {})
        brand_year[r["competitor"]][r["_year"]] += 1
        for t in tags.get("themes", []):
            theme_year[t][r["_year"]] += 1

    def pct(theme, year):
        n = per_year[year]
        return round(100 * theme_year[theme][year] / n) if n else 0

    wb = Workbook()

    ws = wb.active
    ws.title = "Trend"
    ws.append(["Theme"] + [str(y) for y in years] + ["Change"])
    ws.append(["Ads in sample"] + [per_year[y] for y in years] + [""])
    ordered = sorted(THEMES, key=lambda t: -sum(theme_year[t].values()))
    for t in ordered:
        first, last = pct(t, years[0]), pct(t, years[-1])
        ws.append([t] + [pct(t, y) for y in years] + [last - first])
    for row in ws.iter_rows(min_row=1, max_row=1):
        for c in row:
            c.fill, c.font = HEAD_FILL, Font(bold=True, color="FFFFFFFF")
    for row in ws.iter_rows(min_row=3, min_col=2, max_col=len(years) + 1):
        for c in row:
            c.number_format = '0"%"'
            c.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 32
    for i in range(2, len(years) + 3):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.freeze_panes = "B3"

    ws2 = wb.create_sheet("Sample by brand")
    ws2.append(["Brand"] + [str(y) for y in years] + ["Total"])
    for b in sorted(brand_year, key=lambda x: -sum(brand_year[x].values())):
        ws2.append([b] + [brand_year[b][y] for y in years] + [sum(brand_year[b].values())])
    for c in ws2[1]:
        c.fill, c.font = HEAD_FILL, Font(bold=True, color="FFFFFFFF")
    ws2.column_dimensions["A"].width = 22

    ws3 = wb.create_sheet("Tagged Ads")
    ws3.append(["Year", "Competitor", "Ad copy", "Themes", "Audiences", "Proof points"])
    for r in sorted(rows, key=lambda x: (x["_year"], x["competitor"])):
        t = cache.get(r["_key"], {})
        ws3.append([
            r["_year"], r["competitor"], r["ad_text"],
            ", ".join(t.get("themes", [])),
            ", ".join(t.get("audiences", [])),
            "; ".join(t.get("proof_points", [])),
        ])
    for c in ws3[1]:
        c.fill, c.font = HEAD_FILL, Font(bold=True, color="FFFFFFFF")
    ws3.column_dimensions["C"].width = 80
    ws3.freeze_panes = "A2"

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    return years, per_year, theme_year, ordered, pct


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    retag = "--retag" in sys.argv
    path = args[0] if args else os.path.join(BASE, "data", "competitor_ads_5yr_v2.csv")

    rows = load_rows(path)
    print(f"Read {len(rows)} dated ads from {os.path.relpath(path, BASE)}.")
    cache = tag_all(rows, retag)
    years, per_year, theme_year, ordered, pct = build(rows, cache)

    print(f"\nSaved to {os.path.relpath(OUT_PATH, BASE)}\n")
    width = max(len(t) for t in THEMES) + 2
    print("THEME".ljust(width) + "".join(f"{y:>7}" for y in years) + "   change")
    print("(ads)".ljust(width) + "".join(f"{per_year[y]:>7}" for y in years))
    print("-" * (width + 7 * len(years) + 9))
    for t in ordered:
        first, last = pct(t, years[0]), pct(t, years[-1])
        print(t.ljust(width) + "".join(f"{pct(t, y):>6}%" for y in years)
              + f"{last - first:>+8}")


if __name__ == "__main__":
    main()
